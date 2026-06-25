"""Locally-hosted Credit Card Analytics Dashboard & Ledger.

Run:  python3 app.py    then open http://127.0.0.1:5000
All processing is local; PDFs are parsed in-memory and never leave the machine.
"""
import io
import datetime as _dt
import traceback

from flask import Flask, request, jsonify, send_from_directory, Response

import db
import parsers

app = Flask(__name__, static_folder="static", static_url_path="")
app.config["MAX_CONTENT_LENGTH"] = 40 * 1024 * 1024  # 40 MB per request

db.init_db()


@app.route("/")
def index():
    return send_from_directory("static", "index.html")


@app.post("/api/import")
def api_import():
    password = request.form.get("password") or None
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No files uploaded"}), 400

    results = []
    for f in files:
        name = f.filename or "statement.pdf"
        try:
            data = f.read()
            parsed = parsers.parse_statement(io.BytesIO(data), password=password,
                                             file_name=name)
            if not parsed.get("transactions"):
                results.append({"file": name, "ok": False,
                                "error": "No transactions found — wrong password or unsupported layout."})
                continue
            res = db.import_statement(parsed)
            res["file"] = name
            res["ok"] = True
            res["card_label"] = parsed.get("card_label")
            results.append(res)
        except Exception as e:  # noqa: BLE001 — surface parse/decrypt errors to UI
            msg = str(e).strip()
            ename = e.__class__.__name__
            blob = f"{ename} {msg} {e!r}".lower()
            # pdfplumber raises an empty-message PdfminerException when a
            # password-protected PDF can't be decrypted — give a clear hint.
            if (not msg and ename == "PdfminerException") or any(
                    k in blob for k in ("password", "decrypt", "crypt")):
                msg = ("Could not open this PDF — it looks encrypted. Enter the "
                       "exact statement password. Tip: if you're importing "
                       "statements from different cards, import them one at a "
                       "time, since each can have its own password.")
            elif not msg:
                msg = ename
            results.append({"file": name, "ok": False, "error": msg,
                            "trace": traceback.format_exc().splitlines()[-1]})
    return jsonify({"results": results})


def _filters_from_args(src):
    keys = ("search", "category", "subcategory", "bank", "card_last4", "direction",
            "section", "is_emi", "month", "date_from", "date_to", "sort", "order")
    return {k: src.get(k) for k in keys if src.get(k)}


@app.get("/api/transactions")
def api_transactions():
    return jsonify({"transactions": db.list_transactions(_filters_from_args(request.args))})


@app.post("/api/transactions/<int:txn_id>/category")
def api_set_category(txn_id):
    body = request.get_json(force=True, silent=True) or {}
    cat = body.get("category")
    if not cat:
        return jsonify({"error": "category required"}), 400
    return jsonify({"updated": db.update_category(txn_id, cat)})


@app.post("/api/transactions/<int:txn_id>/subcategory")
def api_set_subcategory(txn_id):
    body = request.get_json(force=True, silent=True) or {}
    return jsonify({"updated": db.update_subcategory(txn_id, body.get("subcategory", ""))})


@app.post("/api/transactions/<int:txn_id>/note")
def api_set_note(txn_id):
    body = request.get_json(force=True, silent=True) or {}
    return jsonify({"updated": db.update_note(txn_id, body.get("note", ""))})


@app.post("/api/transactions/bulk-category")
def api_bulk_category():
    body = request.get_json(force=True, silent=True) or {}
    cat = body.get("category") or None
    sub = body.get("subcategory")  # may be "" to clear, or None to leave alone
    if not cat and sub is None:
        return jsonify({"error": "category or subcategory required"}), 400
    filters = _filters_from_args(body.get("filters") or {})
    return jsonify({"updated": db.bulk_update(filters, category=cat, subcategory=sub)})


@app.get("/api/categories")
def api_categories():
    return jsonify({"categories": db.list_categories()})


@app.post("/api/categories")
def api_add_category():
    body = request.get_json(force=True, silent=True) or {}
    name = db.add_category(body.get("name"))
    if not name:
        return jsonify({"error": "name required"}), 400
    return jsonify({"categories": db.list_categories(), "added": name})


@app.get("/api/subcategories")
def api_subcategories():
    return jsonify({"subcategories": db.list_subcategories()})


@app.post("/api/subcategories")
def api_add_subcategory():
    body = request.get_json(force=True, silent=True) or {}
    name = db.add_subcategory(body.get("name"))
    if not name:
        return jsonify({"error": "name required"}), 400
    return jsonify({"subcategories": db.list_subcategories(), "added": name})


@app.get("/api/statements")
def api_statements():
    return jsonify({"statements": db.list_statements()})


@app.delete("/api/statements/<int:stmt_id>")
def api_delete_statement(stmt_id):
    db.delete_statement(stmt_id)
    return jsonify({"deleted": stmt_id})


@app.get("/api/filters")
def api_filters():
    d = db.distinct_values()
    d["all_categories"] = db.list_categories()
    d["all_subcategories"] = db.list_subcategories()
    return jsonify(d)


@app.get("/api/export.xlsx")
def api_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    rows = db.list_transactions(_filters_from_args(request.args))
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"
    headers = ["Date", "Billing cycle", "Bank", "Card", "Description", "Merchant",
               "City", "Category", "Subcategory", "Note", "Section", "Direction",
               "Foreign ccy", "Foreign amt", "Amount (INR)", "EMI", "Reward pts",
               "Reference", "Cardholder"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")
    for t in rows:
        signed = -t["amount"] if t["direction"] == "credit" else t["amount"]
        ws.append([
            t["txn_date"], t.get("cycle_month"), t["bank"], t["card_last4"],
            t["description"], t.get("merchant"), t.get("city"), t.get("category"),
            t.get("subcategory"), t.get("note"), t.get("section"), t["direction"],
            t.get("foreign_currency"), t.get("foreign_amount"), signed,
            "Yes" if t.get("is_emi") else "", t.get("reward_points"),
            str(t.get("ref_no") or ""), t.get("cardholder"),
        ])
    widths = [12, 13, 8, 8, 40, 26, 14, 20, 18, 28, 14, 9, 11, 12, 14, 6, 10, 22, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    amt_col = headers.index("Amount (INR)") + 1
    fa_col = headers.index("Foreign amt") + 1
    for r in range(2, ws.max_row + 1):
        ws.cell(r, amt_col).number_format = '#,##0.00'
        ws.cell(r, fa_col).number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"ledger_{_dt.date.today().isoformat()}.xlsx"
    return Response(buf.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@app.get("/api/stats")
def api_stats():
    return jsonify(db.stats_overview(
        card=request.args.get("card") or None,
        month=request.args.get("month") or None))


@app.get("/api/reconciliation")
def api_reconciliation():
    return jsonify({"reconciliation": db.reconciliation()})


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", "5000"))
    print(f"\n  Credit Card Dashboard running at  http://127.0.0.1:{port}\n")
    app.run(host="127.0.0.1", port=port, debug=False)
